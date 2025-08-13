import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import cm_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D
from io import BytesIO
from PIL import Image as PILImage
import io
import base64
import gspread
from gspread_dataframe import set_with_dataframe
from google.oauth2.service_account import Credentials
import dropbox

# --- Configuración de las dimensiones del área de la imagen (en cm) ---

# -- Función para autenticar y subir archivo a Google Drive
def upload_to_dropbox(file_buffer, filename):
    """
    Sube un archivo a Dropbox usando un token de acceso.
    """
    try:
        # Obtén el token de acceso desde los secretos de Streamlit
        dbx = dropbox.Dropbox(st.secrets['dropbox_access_token'])

        # El path de destino en tu cuenta de Dropbox
        # Por ejemplo, /Apps/NombreDeTuApp/nombre_del_archivo
        dropbox_path = f"/Apps/Streamlit App Uploader/{filename}"

        # Convierte el buffer de archivo a bytes para subirlo
        file_buffer.seek(0)
        data = file_buffer.read()

        # Sube el archivo a Dropbox
        dbx.files_upload(data, dropbox_path, mode=dropbox.files.WriteMode('overwrite'))

        st.success(f"¡El archivo '{filename}' ha sido subido exitosamente a Dropbox!")
        return True

    except dropbox.exceptions.AuthError as err:
        st.error("Error de autenticación. Verifica tu token de acceso en Streamlit secrets.")
        return False
    except Exception as e:
        st.error(f"Ocurrió un error al subir el archivo: {e}")
        return False


# -- Función llenar preventivo/recorredor
def preventivo_recorredor(formato_seleccionado,ejecutor,direccion,fecha_visita,operador,cambio):
    ruta_excel = 'RF_PREVENTIVO.XLSX'
    libro = load_workbook(ruta_excel)
    hoja = libro.active
    AREA_HEIGHT_CM = 6.8
    AREA_WIDTH_CM = 9.42

    fila_foto_inicio = 8
    columna_foto_inicio = 1  # Columna A
    
    hoja['G5'] = ejecutor
    hoja['C8'] = direccion
    hoja['C50'] = fecha_visita.strftime("%d-%m-%Y")
    hoja['G6'] = cambio

    hoja['G8'] = operador

    if formato_seleccionado == "Recorredor":
        hoja['A4'] = 'REGISTRO FOTOGRÁFICO RECORREDOR'
        hoja['D7'] = 'RECORREDOR'

    return fila_foto_inicio,AREA_WIDTH_CM, AREA_HEIGHT_CM,columna_foto_inicio,libro

# -- Función llenar cliente interno/externo
def interno_externo(formato_seleccionado,ejecutor,direccion,operador,cliente,cambio):
    AREA_HEIGHT_CM = 6.8
    AREA_WIDTH_CM = 9.42
    ruta_excel = ''
    
    if formato_seleccionado == "clientes interno" or formato_seleccionado == "Empalmeria":
        ruta_excel = 'RF_CLIENTE_INTERNO.xlsx'
    
    elif formato_seleccionado == "clientes externo":
        ruta_excel = 'RF_CLIENTE_EXTERNO.xlsx'

    libro = load_workbook(ruta_excel)
    hoja = libro.active

    hoja['C6'] = cliente
    fila_foto_inicio = 10
    columna_foto_inicio = 1  # Columna A
    hoja['G5'] = ejecutor
    hoja['C8'] = direccion
    #hoja['C50'] = fecha_visita.strftime("%d-%m-%Y")
    hoja['G8'] = operador
    hoja['G6'] = cambio

    return fila_foto_inicio,AREA_WIDTH_CM, AREA_HEIGHT_CM,columna_foto_inicio,libro


# -- Función llenar factibilidades
def factibilidades(ejecutor, direccion, fecha_visita, cliente, cambio, telefono_ejecutor, encargado, telefono_encargado, atiende_en_sitio, telefono_atiende_sitio):
    # Ajustamos las dimensiones del área de la imagen para que coincidan con las de "interno/externo"
    AREA_HEIGHT_CM = 6.8
    AREA_WIDTH_CM = 9.42
    ruta_excel = 'RF_FACTIBILIDADES.xlsx'
    libro = load_workbook(ruta_excel)
    hoja = libro.active
    
    # Llenamos las celdas del encabezado según las nuevas ubicaciones
    hoja['C5'] = cliente
    hoja['C6'] = direccion
    hoja['C7'] = fecha_visita.strftime("%d-%m-%Y")
    hoja['G5'] = ejecutor
    hoja['H5'] = telefono_ejecutor # Nuevo campo
    hoja['G6'] = encargado        # Nuevo campo
    hoja['H6'] = telefono_encargado # Nuevo campo
    hoja['G7'] = atiende_en_sitio   # Nuevo campo
    hoja['H7'] = telefono_atiende_sitio # Nuevo campo
    hoja['G8'] = cambio
    
    fila_foto_inicio = 12
    columna_foto_inicio = 1 # Columna A
    
    return fila_foto_inicio, AREA_WIDTH_CM, AREA_HEIGHT_CM, columna_foto_inicio, libro


# -- Función llenar cartera
def cartera(ejecutor, direccion, fecha_visita, operador, cliente, archivos_por_poste):
    AREA_HEIGHT_CM = 10.7
    AREA_WIDTH_CM = 12.3
    ruta_excel = 'RF_CARTERA.xlsx'
    libro = load_workbook(ruta_excel)
    hoja1 = libro["CARTERA"]
    hoja2 = libro["REGISTRO FOTOGRAFICO"]
    # hoja = libro.active
    fila_foto_inicio = 2
    columna_foto_inicio = 4  # Inicial en D

    hoja1['B9'] = ejecutor
    hoja1['B7'] = direccion
    hoja1['B8'] = fecha_visita.strftime("%d-%m-%Y")
    hoja1['D9'] = operador

    fila_actual_foto = fila_foto_inicio
    for i, archivos in enumerate(archivos_por_poste):
        col_offset = i * 6  # Espacio para 5 fotos y un espacio
        for j, archivo_subido in enumerate(archivos):
            if archivo_subido:
                img_pil_original = PILImage.open(archivo_subido)
                img_redimensionada = redimensionar_imagen(img_pil_original, AREA_WIDTH_CM, AREA_HEIGHT_CM)
                img_width_cm, img_height_cm = img_redimensionada.size[0] * 2.54 / 96, img_redimensionada.size[1] * 2.54 / 96
                img_buffer = BytesIO()
                img_redimensionada.save(img_buffer, format="PNG")
                img_buffer.seek(0)
                img = Image(img_buffer)

                col_idx = columna_foto_inicio - 1 + col_offset + j
                row_idx = fila_actual_foto - 1

                x_offset_emu = calcular_offset(AREA_WIDTH_CM, img_width_cm)
                y_offset_emu = calcular_offset(AREA_HEIGHT_CM, img_height_cm)

                marker = AnchorMarker(col=col_idx, colOff=x_offset_emu, row=row_idx, rowOff=y_offset_emu)
                size = XDRPositiveSize2D(cx=c2e(img_width_cm), cy=c2e(img_height_cm))
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                hoja2.add_image(img)
        fila_actual_foto += 15  # Espacio para el siguiente poste

    return fila_foto_inicio, AREA_WIDTH_CM, AREA_HEIGHT_CM, columna_foto_inicio, libro

# ---  Función rotar imagen
def rotate_image(image_bytes, angle):
    """Rota una imagen en bytes y la devuelve en bytes."""
    img = PILImage.open(io.BytesIO(image_bytes))
    rotated_img = img.rotate(angle, expand=True)
    img_buffer = io.BytesIO()
    rotated_img.save(img_buffer, format=img.format)
    return img_buffer.getvalue()

# --- Función para redimensionar la imagen manteniendo la relación de aspecto ---
def redimensionar_imagen(imagen_pil, max_ancho_cm, max_alto_cm, dpi=96):
    """Redimensiona una imagen de Pillow manteniendo su relación de aspecto."""
    max_ancho_pixels = max_ancho_cm * dpi / 2.54
    max_alto_pixels = max_alto_cm * dpi / 2.54
    ancho, alto = imagen_pil.size

    ratio_ancho = max_ancho_pixels / ancho
    ratio_alto = max_alto_pixels / alto

    if ratio_ancho < 1 or ratio_alto < 1:  # Solo redimensionar si es más grande
        ratio = min(ratio_ancho, ratio_alto)
        nuevo_ancho = int(ancho * ratio)
        nuevo_alto = int(alto * ratio)
        imagen_redimensionada = imagen_pil.resize((nuevo_ancho, nuevo_alto))
        return imagen_redimensionada
    return imagen_pil

# --- Función para convertir cm a EMUs ---
c2e = cm_to_EMU

# --- Función para calcular el offset en EMUs para centrar la imagen ---
def calcular_offset(area_cm, img_cm):
    delta_cm = ((area_cm - img_cm) / 2)+0.1
    return c2e(delta_cm)

# --- Interfaz de Usuario en Streamlit ---
st.title("Registro fotografico")

formato = ["clientes interno", "clientes externo","Empalmeria","Factibilidades" ]
formato_seleccionado = st.radio("Selecciona el formato:", formato)


opciones = ["DIEGO ARMANDO CHATEZ MARTINEZ","HAROLD ANDRES TORRES TEPUD","VICTOR ANDRES BOTINA ALVAREZ","CARLOS ANDRES MARCILLO","OMAR ALEXANDER DULCE LOPEZ",
            "YESID ALFONSO SANCHEZ DIAZ","ALDIVEY QUINAYAS MUÑOZ","DANIEL EDUARDO TROCHEZ MUÑOZ","ANDRES CAMILO ALEGRIA ALEGRIA","VICTOR ALIRIO ARDILA CELIS",
            "NASPIRAN ROSERO SEGUNDO JUBENAL","MARINO SANCHEZ GARCIA","DIEGO ARMANDO MUÑOZ SAAVEDRA","DIEGO ALEJANDRO VEGA GALEANO","RUTBEL TRUJILLO","VICTOR ALFONSO MORA"
            ]
operadores = {"LIBERTY NETWORK","CLARO","MOVISTAR","TIGO", "IFX NETWOKS","ETB","VERIZON"
            }

ejecutor = st.selectbox("Ejecutor:", opciones)

if formato_seleccionado == "Factibilidades":
    telefono_ejecutor = st.text_input("Teléfono del Ejecutor:", key="telefono_ejecutor")
    encargado = st.text_input("Encargado:")
    telefono_encargado = st.text_input("Teléfono del Encargado:", key="telefono_encargado")
    atiende_en_sitio = st.text_input("Atiende en Sitio:")
    telefono_atiende_sitio = st.text_input("Teléfono de Quien Atiende en Sitio:", key="telefono_atiende_sitio")
    cliente = st.text_input("Nombre del sitio:", key="cliente_factibilidades")
    direccion = st.text_input("DIRECCIÓN:", key="direccion_factibilidades")
    cambio = st.text_input("CAMBIO,TICKET,OT:", key="cambio_factibilidades")
    fecha_visita = st.date_input("FECHA DE LA VISITA:", key="fecha_factibilidades")
    # Se añade un selectbox para el operador, aunque no se usa en la función de llenado, se deja para la interfaz.
    operador = st.selectbox("OPERADOR:", operadores, key="operador_factibilidades_selectbox")
else:
    # Campos para otros formatos
    if formato_seleccionado == "clientes interno" or formato_seleccionado == "clientes externo" or formato_seleccionado == "Empalmeria":
        cliente = st.text_input("Nombre del sitio:", key="cliente_otros")
    else:
        cliente = "" # Para evitar que la variable no esté definida
        
    direccion = st.text_input("DIRECCIÓN:", key="direccion_otros")
    cambio = st.text_input("CAMBIO,TICKET,OT:", key="cambio_otros")
    fecha_visita = st.date_input("FECHA DE LA VISITA:", key="fecha_otros")
    operador = st.selectbox("OPERADOR:", operadores, key="operador_otros_selectbox")
    telefono_ejecutor = ""
    encargado = ""
    telefono_encargado = ""
    atiende_en_sitio = ""
    telefono_atiende_sitio = ""


archivos_por_poste = []



#diligenciar cartera
if formato_seleccionado == "Cartera":
    
    if 'num_postes' in st.session_state and st.session_state['num_postes'] > 0:
        st.subheader("Cargar Fotos por Poste")
        for i in range(st.session_state['num_postes']):
            st.write(f"**Poste {i + 1}:**")
            
            fotos_poste = {}
            col_uploaders = st.columns(1)
            fotos_poste["DIRECCION"] = st.file_uploader("DIRECCION", type=["png", "jpg", "jpeg"], key=f"direccion_poste_{i}")
            fotos_poste["ESTADO"] = st.file_uploader("ESTADO", type=["png", "jpg", "jpeg"], key=f"estado_poste_{i}")
            fotos_poste["VISTA GENERAL"] = st.file_uploader("VISTA GENERAL", type=["png", "jpg", "jpeg"], key=f"vista_general_{i}")
            fotos_poste["SENTIDO VIA"] = st.file_uploader("SENTIDO VIA", type=["png", "jpg", "jpeg"], key=f"sentido_via_{i}")
            fotos_poste["VISTA GENERAL SIGUIENTE INFRAESTRUCTURA"] = st.file_uploader(
                "VISTA GENERAL SIGUIENTE INFRAESTRUCTURA", type=["png", "jpg", "jpeg"], key=f"vista_general_sig_{i}")
            archivos_por_poste.append({"fotos": fotos_poste})
            
            st.subheader(f"Vista Previa del Poste {i + 1}")
            col_preview = st.columns(5)
            nombres_campos = ["DIRECCION", "ESTADO", "VISTA GENERAL", "SENTIDO VIA", "VISTA GENERAL SIGUIENTE INFRAESTRUCTURA"]
            for j, nombre_campo in enumerate(nombres_campos):
                uploaded_file = fotos_poste.get(nombre_campo)
                if uploaded_file:
                    try:
                        img = PILImage.open(uploaded_file)
                        col_preview[j].image(img, caption=nombre_campo, width=100)
                    except Exception as e:
                        col_preview[j].error(f"Error al cargar {nombre_campo}: {e}")

    if st.button("Agregar Poste"):
        st.session_state['num_postes'] = st.session_state.get('num_postes', 0) + 1
        st.rerun()

else:
    uploaded_files = st.file_uploader("Subir Registros Fotográficos", accept_multiple_files=True,
                                     type=["png", "jpg", "jpeg"])

    descripciones = [""] * len(uploaded_files)

    # --- Sección de Vista Previa con Miniaturas y Campos de Descripción Adyacentes ---
    if ejecutor or direccion or fecha_visita or operador or uploaded_files:
        st.subheader("Vista Previa de los Datos:")
        if ejecutor:
            st.write(f"**Ejecutor:** {ejecutor}")
        if direccion:
            st.write(f"**Dirección:** {direccion}")
        if fecha_visita:
            st.write(f"**Fecha de la Visita:** {fecha_visita.strftime('%Y-%m-%d')}")
        if operador:
            st.write(f"**Teléfono:** {operador}")
        if uploaded_files and formato_seleccionado != "Cartera":
            st.write("**Registros Fotográficos:**")
            if formato_seleccionado == "Factibilidades":
                num_filas_preview = (len(uploaded_files) + 2) // 3
                for i in range(num_filas_preview):
                    cols = st.columns(3)
                    for j in range(3):
                        idx = i * 3 + j
                        if idx < len(uploaded_files):
                            file = uploaded_files[idx]
                            key_rotacion = f"rotacion_{idx}"
                            key_imagen_rotada = f"imagen_rotada_{idx}"

                            if key_rotacion not in st.session_state:
                                st.session_state[key_rotacion] = 0
                            
                            with cols[j]:
                                col_imagen_botones = st.columns([3, 2])
                                with col_imagen_botones[0]:
                                    try:
                                        img = PILImage.open(file)
                                        if st.session_state[key_rotacion] != 0:
                                            rotated_img = img.rotate(st.session_state[key_rotacion], expand=True)
                                            st.image(rotated_img,
                                                    caption=f"Foto {idx + 1} (Rotada {st.session_state[key_rotacion]}°)",
                                                    width=100)
                                        else:
                                            st.image(img, caption=f"Foto {idx + 1}", width=100)
                                    except Exception as e:
                                        st.error(f"Error: No se pudo abrir el archivo como imagen: {file.name}")

                                with col_imagen_botones[1]:
                                    col_rot_left, col_rot_right = st.columns(2)
                                    with col_rot_left:
                                        if st.button("↺", key=f"rotar_der_{idx}"):
                                            st.session_state[key_rotacion] = (st.session_state[key_rotacion] + 90) % 360
                                            st.rerun()

                                    with col_rot_right:
                                        if st.button("↻", key=f"rotar_izq_{idx}"):
                                            st.session_state[key_rotacion] = (st.session_state[key_rotacion] - 90) % 360
                                            st.rerun()
                                descripcion_key = f"descripcion_factibilidad_{i}"
                                descripciones.extend([""] * 3)
                                descripciones[i * 3: (i + 1) * 3] = [st.text_input(
                                    f"Descripción para Fotos {(i * 3) + 1} a {(i + 1) * 3}:", key=descripcion_key)] * 3
            else:
                for i, file in enumerate(uploaded_files):
                    key_rotacion = f"rotacion_{i}"
                    key_imagen_rotada = f"imagen_rotada_{i}"

                    if key_rotacion not in st.session_state:
                        st.session_state[key_rotacion] = 0

                    col_imagen_botones = st.columns([3, 1])
                    with col_imagen_botones[0]:
                        try:
                            img = PILImage.open(file)
                            if st.session_state[key_rotacion] != 0:
                                rotated_img = img.rotate(st.session_state[key_rotacion], expand=True)
                                st.image(rotated_img, caption=f"Foto {i+1} (Rotada {st.session_state[key_rotacion]}°)", width=100)
                            else:
                                st.image(img, caption=f"Foto {i+1}", width=100)
                        except Exception as e:
                            st.error(f"Error: No se pudo abrir el archivo como imagen: {file.name}")

                    with col_imagen_botones[1]:
                        col_rot_left, col_rot_right = st.columns(2)
                        with col_rot_left:
                            if st.button("↺", key=f"rotar_der_{i}"):
                                st.session_state[key_rotacion] = (st.session_state[key_rotacion] + 90) % 360
                                st.rerun()
                        
                        with col_rot_right:
                            if st.button("↻", key=f"rotar_izq_{i}"):
                                st.session_state[key_rotacion] = (st.session_state[key_rotacion] - 90) % 360
                                st.rerun()
                        
                    descripciones[i] = st.text_input(f"Descripción para la Foto {i+1}:", key=f"descripcion_{i}", value= file.name)
            


if st.button("Enviar a Drive"):
    if ejecutor and direccion and fecha_visita and operador and uploaded_files:
        try:
            if formato_seleccionado == "Preventivo":
                fila_foto_inicio,AREA_WIDTH_CM, AREA_HEIGHT_CM,columna_foto_inicio,libro = preventivo_recorredor(formato_seleccionado,ejecutor,direccion,fecha_visita,operador,cambio)
            elif formato_seleccionado == "Recorredor":
                fila_foto_inicio,AREA_WIDTH_CM, AREA_HEIGHT_CM,columna_foto_inicio,libro = preventivo_recorredor(formato_seleccionado,ejecutor,direccion,fecha_visita,operador,cambio)
            elif formato_seleccionado == "clientes interno" or formato_seleccionado == "Empalmeria":
                fila_foto_inicio,AREA_WIDTH_CM, AREA_HEIGHT_CM,columna_foto_inicio,libro = interno_externo(formato_seleccionado,ejecutor,direccion,operador,cliente,cambio)
            elif formato_seleccionado == "clientes externo":
                fila_foto_inicio,AREA_WIDTH_CM, AREA_HEIGHT_CM,columna_foto_inicio,libro = interno_externo(formato_seleccionado,ejecutor,direccion,operador,cliente,cambio)
            elif formato_seleccionado == "Factibilidades":
                fila_foto_inicio,AREA_WIDTH_CM, AREA_HEIGHT_CM,columna_foto_inicio,libro = factibilidades(ejecutor, direccion, fecha_visita, cliente, cambio, telefono_ejecutor, encargado, telefono_encargado, atiende_en_sitio, telefono_atiende_sitio)
            elif formato_seleccionado == "Cartera":
                fila_foto_inicio,AREA_WIDTH_CM, AREA_HEIGHT_CM,columna_foto_inicio,libro = cartera(ejecutor,direccion,fecha_visita,operador,cliente,archivos_por_poste)
            
            hoja = libro.active
            print(libro)
            
            fila_actual_foto = fila_foto_inicio
            for i, archivo_subido in enumerate(uploaded_files):
                key_rotacion = f"rotacion_{i}"
                angulo_rotacion = st.session_state.get(key_rotacion, 0) # Obtener el ángulo de rotación

                # --- Rotar la imagen antes de redimensionar ---
                img_pil_original = PILImage.open(archivo_subido)
                img_pil_rotada = img_pil_original.rotate(angulo_rotacion, expand=True)


                # --- Redimensionar la imagen ---
                #img_pil = PILImage.open(archivo_subido)
                img_redimensionada = redimensionar_imagen(img_pil_rotada, AREA_WIDTH_CM, AREA_HEIGHT_CM)
                img_width_cm, img_height_cm = img_redimensionada.size[0] * 2.54 / 96, img_redimensionada.size[1] * 2.54 / 96

                # --- Convertir la imagen para openpyxl ---
                img_buffer = BytesIO()
                img_redimensionada.save(img_buffer, format="PNG")
                img_buffer.seek(0)
                img = Image(img_buffer)

                # --- Calcular la columna de anclaje para Factibilidades ---
                if formato_seleccionado == "Factibilidades":
                    if i % 3 == 0:
                        col_idx = 0  # Columna A
                    elif i % 3 == 1:
                        col_idx = 3  # Columna D
                    else:
                        col_idx = 6  # Columna G
                else: # Para otros formatos, la lógica original
                    col_idx = (columna_foto_inicio - 1) + (4 * (i % 2))

                row_idx = fila_actual_foto - 1

                x_offset_emu = calcular_offset(AREA_WIDTH_CM, img_width_cm)
                y_offset_emu = calcular_offset(AREA_HEIGHT_CM, img_height_cm)

                # --- Definir el marcador de anclaje ---
                marker = AnchorMarker(col=col_idx, colOff=x_offset_emu, row=row_idx, rowOff=y_offset_emu)
                size = XDRPositiveSize2D(cx=c2e(img_width_cm), cy=c2e(img_height_cm))
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                hoja.add_image(img)

                # --- Calcular la celda unificada para la descripción para Factibilidades ---
                if formato_seleccionado == "Factibilidades":
                    celda_descripcion_inicio = f"B{fila_actual_foto + 1}"
                    celda_descripcion_fin = f"H{fila_actual_foto + 2}"
                    hoja.merge_cells(f"{celda_descripcion_inicio}:{celda_descripcion_fin}")
                    hoja[celda_descripcion_inicio] = descripciones[i]
                    if (i + 1) % 3 == 0: # Saltar cada 3 imágenes (al final de la fila A, D, G)
                        fila_actual_foto += 6
                else: # Lógica de descripción para otros formatos
                    if (i + 1) % 2 != 0:
                        celda_descripcion_inicio = f"B{fila_actual_foto + 15}"
                        celda_descripcion_fin = f"D{fila_actual_foto + 16}"
                    else:
                        celda_descripcion_inicio = f"F{fila_actual_foto + 15}"
                        celda_descripcion_fin = f"H{fila_actual_foto + 16}"
                        fila_actual_foto += 17
                    hoja.merge_cells(f"{celda_descripcion_inicio}:{celda_descripcion_fin}")
                    hoja[celda_descripcion_inicio] = descripciones[i]

            # --- Guardar y subir a Drive ---
            buffer = BytesIO()
            libro.save(buffer)
            buffer.seek(0)

            # Lógica para determinar el acrónimo del formato
            if formato_seleccionado == "clientes interno":
                formato_acronimo = "int"
                filename = f"{fecha_visita.strftime('%d-%m-%Y')}_{cambio}_{cliente}_{formato_acronimo}.xlsx"
            elif formato_seleccionado == "clientes externo":
                formato_acronimo = "ext"
                filename = f"{fecha_visita.strftime('%d-%m-%Y')}_{cambio}_{cliente}_{formato_acronimo}.xlsx"
            elif formato_seleccionado == "Empalmeria":
                formato_acronimo = "emp"
                filename = f"{fecha_visita.strftime('%d-%m-%Y')}_{cambio}_{cliente}_{formato_acronimo}.xlsx"
            elif formato_seleccionado == "Factibilidades":
                formato_acronimo = "fac"
                filename = f"{fecha_visita.strftime('%d-%m-%Y')}_{cambio}_{cliente}_{formato_acronimo}.xlsx"
            else:
                filename = f"{fecha_visita.strftime('%d-%m-%Y')}_{cambio}.xlsx"
            
            upload_to_dropbox(buffer, filename)
            
            st.success(f"¡El archivo '{filename}' ha sido subido exitosamente a Dropbox!")

        except Exception as e:
            st.error(f"Ocurrió un error: {e}")
    else:
        st.warning("Por favor, completa todos los campos y sube al menos una foto.")